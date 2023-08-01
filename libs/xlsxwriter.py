import pandas as pd  # type: ignore
from openpyxl.styles import Alignment  # type: ignore
from openpyxl.styles import PatternFill  # type: ignore

from libs.hasic import Hasic


class Xlsxwriter:
    def __init__(self, hasici: list[Hasic]) -> None:
        self.hasici = hasici

    def write(self, soubor: str) -> None:
        while True:
            try:
                writer = pd.ExcelWriter(soubor, engine='openpyxl')
                break
            except PermissionError:
                print(
                    f'Soubor {soubor} je používán jinou aplikací.' +
                    'Zavřete soubor a stiskněte Enter.',
                )
                input()
        data = {}
        for hasic in self.hasici:
            zaznamy = {}
            for vyz in hasic.vyznamenani:
                zaznamy[vyz.nazev] = ', '.join(vyz.splnene_podminky(hasic))
            data[hasic.jmeno] = zaznamy
        df = pd.DataFrame(data)
        transposed = df.T
        transposed.to_excel(
            writer,
            sheet_name='splnene podminky',
            header=True,
            index=True,
            # encoding='utf8',
        )
        ws = writer.sheets['splnene podminky']
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                _cell = ws.cell(column=col, row=row)
                _cell.alignment = Alignment(wrap_text=True)
                if _cell.value:
                    if (
                        'Obdrženo' in _cell.value or
                        'Již se nedává' in _cell.value
                    ):
                        _cell.fill = PatternFill(
                            start_color='00969696',
                            fill_type='solid',
                        )
                    elif 'Podmínky splněny' in _cell.value:
                        _cell.fill = PatternFill(
                            start_color='00CCFFCC',
                            fill_type='solid',
                        )
        ws.freeze_panes = ws['A2']
        writer.close()
